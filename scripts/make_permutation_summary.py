"""
scripts/make_permutation_summary.py

Build comparison tables for the permutation (random-baseline) ablation study.

Reads
-----
    data/processed/backtests/permutation_test/
        {method}_ew_inter_markowitz_intra_k{k}_seed{seed:04d}/
            ..._daily_portfolio_returns.csv
            ..._monthly_portfolio_returns.csv

    data/processed/backtests/{method}_ew_inter_markowitz_intra/
        {method}_ew_inter_markowitz_intra_k{k}/
            ..._daily_portfolio_returns.csv
            ..._monthly_portfolio_returns.csv

Writes
------
    data/processed/backtests/permutation_summary/ew_inter_markowitz_intra/
        k{k}/
            kmeans_comparison/
                summary_metrics.csv
                summary_metrics.xlsx      ← green=original, grey=seeds, orange=random_mean
                monthly_returns_comparison.csv
                rolling_sharpe_comparison.csv
            kmedoids_comparison/
                (same 4 files)
        all_k/
            kmeans/
                value_added.csv/.xlsx     ← per-k original vs random_mean delta
                k_comparison.csv/.xlsx    ← all k ordered: k1_original, k1_random_mean, ...
            kmedoids/
                (same 4 files)

CLI
---
    python scripts/make_permutation_summary.py
    python scripts/make_permutation_summary.py --method kmeans
    python scripts/make_permutation_summary.py --k_values 3 5 10
"""
from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path
import sys

import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.metrics.performance import compute_performance_metrics
from src.paths import get_backtests_dir, UNIVERSE_CHOICES

# ── Paths — overridden inside main() via global ────────────────────────────────

PERM_TEST_ROOT   = PROJECT_ROOT / "data" / "processed" / "backtests" / "permutation_test"
BACKTESTS_ROOT   = PROJECT_ROOT / "data" / "processed" / "backtests"
SUMMARY_ROOT     = PROJECT_ROOT / "data" / "processed" / "backtests" / "summary" / "permutation_summary" / "ew_inter_markowitz_intra"

ROLLING_WINDOW   = 12
RISK_FREE_RATE   = 0.0

_METHOD_PREFIX = {
    "kmeans":   "kmeans_ew_inter_markowitz_intra",
    "kmedoids": "kmedoids_dtw_ew_inter_markowitz_intra",
}

# ── Colours for xlsx ──────────────────────────────────────────────────────────

_COLOR_ORIGINAL    = "E2EFDA"   # green
_COLOR_SEED        = "F2F2F2"   # light grey
_COLOR_RANDOM_MEAN = "FCE4D6"   # orange

# ── Data loading ──────────────────────────────────────────────────────────────

def _load_daily(csv_path: Path) -> pd.Series:
    """Load a daily portfolio-return CSV as a float64 Series.

    Parameters
    ----------
    csv_path : Path
        Path to a strategy CSV with a date index and a
        ``portfolio_return`` column (or any single numeric column).

    Returns
    -------
    pd.Series
        Float64 return series sorted by date, named
        ``"portfolio_return"``.
    """
    df = pd.read_csv(csv_path, index_col=0, parse_dates=True)
    col = "portfolio_return" if "portfolio_return" in df.columns else df.columns[0]
    s = pd.to_numeric(df[col], errors="coerce").sort_index().astype("float64")
    s.name = "portfolio_return"
    return s


def _load_monthly(csv_path: Path) -> pd.Series:
    """Load a monthly portfolio-return CSV as a float64 Series.

    Parameters
    ----------
    csv_path : Path
        Path to a monthly return CSV produced by the backtest pipeline.

    Returns
    -------
    pd.Series
        Float64 monthly return series sorted by date.
    """
    return _load_daily(csv_path)


def _monthly_from_daily(daily: pd.Series) -> pd.Series:
    """Compound daily returns into month-end monthly returns.

    Parameters
    ----------
    daily : pd.Series
        Daily return series with a DatetimeIndex.

    Returns
    -------
    pd.Series
        Month-end compounded returns as a float64 Series.
    """
    return ((1 + daily).resample("ME").prod() - 1).astype("float64")


def _find_file(directory: Path, suffix: str) -> Path | None:
    """Find a CSV file ending with *suffix* inside directory.

    Parameters
    ----------
    directory : Path
        Directory to search for matching CSV files.
    suffix : str
        Filename suffix to match (e.g. ``"_daily_portfolio_returns"``).

    Returns
    -------
    Path or None
        First alphabetically matching ``.csv`` file, or ``None`` if
        no match is found.
    """
    matches = sorted(directory.glob(f"*{suffix}.csv"))
    return matches[0] if matches else None


def _load_series_from_dir(directory: Path) -> tuple[pd.Series | None, pd.Series | None]:
    """Return (daily, monthly) from a strategy output directory.

    Parameters
    ----------
    directory : Path
        Strategy output directory containing daily and/or monthly
        portfolio-return CSV files.

    Returns
    -------
    tuple[pd.Series or None, pd.Series or None]
        ``(daily, monthly)`` return series. If the monthly CSV is
        missing it is derived by compounding the daily series; if both
        are absent both elements are ``None``.
    """
    daily_path   = _find_file(directory, "_daily_portfolio_returns")
    monthly_path = _find_file(directory, "_monthly_portfolio_returns")

    daily   = _load_daily(daily_path)     if daily_path   else None
    monthly = _load_monthly(monthly_path) if monthly_path else None

    if monthly is None and daily is not None:
        monthly = _monthly_from_daily(daily)

    return daily, monthly


# ── Sortino helper ────────────────────────────────────────────────────────────

def _sortino(monthly: pd.Series, risk_free_rate: float = 0.0) -> float:
    """Annualized Sortino ratio from monthly returns.

    Parameters
    ----------
    monthly : pd.Series
        Monthly portfolio return series.
    risk_free_rate : float, optional
        Annualized risk-free rate; divided by 12 internally.

    Returns
    -------
    float
        Annualized Sortino ratio, or ``NaN`` if there are fewer than
        two negative-excess-return observations or downside std is zero.
    """
    rf_monthly  = risk_free_rate / 12.0
    excess      = monthly - rf_monthly
    downside    = excess[excess < 0]
    if len(downside) < 2:
        return np.nan
    downside_std = float(downside.std(ddof=1))
    if np.isclose(downside_std, 0.0):
        return np.nan
    ann_excess   = float(excess.mean()) * 12.0
    ann_down_std = downside_std * np.sqrt(12.0)
    return ann_excess / ann_down_std


# ── Metrics computation ───────────────────────────────────────────────────────

def _compute_metrics(daily: pd.Series | None, monthly: pd.Series) -> dict:
    """Compute performance and Sortino metrics from daily or monthly returns.

    Parameters
    ----------
    daily : pd.Series or None
        Daily return series used when available (preferred for accuracy);
        if ``None`` the monthly series is used instead.
    monthly : pd.Series
        Monthly return series used for the Sortino ratio calculation and
        as a fallback for other metrics when ``daily`` is ``None``.

    Returns
    -------
    dict
        Keys: ``annualized_return``, ``annualized_volatility``,
        ``sharpe_ratio``, ``sortino_ratio``, ``max_drawdown``,
        ``calmar_ratio``, ``start_date``, ``end_date``.
    """
    if daily is not None:
        base = compute_performance_metrics(
            daily, periods_per_year=252, risk_free_rate=RISK_FREE_RATE
        )
    else:
        base = compute_performance_metrics(
            monthly, periods_per_year=12, risk_free_rate=RISK_FREE_RATE
        )

    sortino = _sortino(monthly, RISK_FREE_RATE)
    return {
        "annualized_return":     base["annualized_return"],
        "annualized_volatility": base["annualized_volatility"],
        "sharpe_ratio":          base["sharpe_ratio"],
        "sortino_ratio":         sortino,
        "max_drawdown":          base["max_drawdown"],
        "calmar_ratio":          base["calmar_ratio"],
        "start_date":            base["start_date"],
        "end_date":              base["end_date"],
    }


def _metrics_row(label: str, daily: pd.Series | None, monthly: pd.Series) -> dict:
    """Build one named summary-metrics row.

    Parameters
    ----------
    label : str
        Strategy name used as the ``"strategy"`` key in the returned dict.
    daily : pd.Series or None
        Daily return series (passed through to ``_compute_metrics``).
    monthly : pd.Series
        Monthly return series (passed through to ``_compute_metrics``).

    Returns
    -------
    dict
        Metrics dict with a ``"strategy"`` key prepended.
    """
    m = _compute_metrics(daily, monthly)
    return {"strategy": label, **m}


# ── Rolling Sharpe ────────────────────────────────────────────────────────────

def _rolling_sharpe(monthly_df: pd.DataFrame, window: int = 12) -> pd.DataFrame:
    """Compute annualized rolling Sharpe ratios from monthly returns.

    Parameters
    ----------
    monthly_df : pd.DataFrame
        Wide-format monthly return DataFrame, columns = strategy names.
    window : int, optional
        Rolling window length in months.

    Returns
    -------
    pd.DataFrame
        Annualized rolling Sharpe ratios with the same shape and index as
        ``monthly_df``; columns sorted alphabetically.
    """
    if monthly_df.empty:
        return pd.DataFrame()
    rm  = monthly_df.rolling(window=window).mean()
    rs  = monthly_df.rolling(window=window).std(ddof=1)
    ann = rm * 12.0 / (rs * np.sqrt(12.0))
    ann = ann.replace([np.inf, -np.inf], np.nan)
    ann.index.name = "date"
    return ann.sort_index(axis=1)


# ── Excel saving ──────────────────────────────────────────────────────────────

def _save_xlsx_colored(df: pd.DataFrame, path: Path, row_colors: dict[str, str]) -> None:
    """Save DataFrame to xlsx; apply per-row background fill based on row_colors[strategy_name].

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to write; must contain a ``"strategy"`` column for
        color lookup.
    path : Path
        Destination ``.xlsx`` file path (parent directory is created if
        needed).
    row_colors : dict[str, str]
        Mapping from strategy name to a six-character hex fill color
        (without ``#``).
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, sheet_name="summary")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print(f"[WARN] openpyxl not available — skipping color formatting: {path}")
        return

    wb  = load_workbook(path)
    ws  = wb.active

    header = [cell.value for cell in ws[1]]
    strat_col = (header.index("strategy") + 1) if "strategy" in header else None

    for row_idx in range(2, ws.max_row + 1):
        strat_val = ws.cell(row=row_idx, column=strat_col).value if strat_col else None
        color     = row_colors.get(str(strat_val) if strat_val else "", None)
        if color:
            fill = PatternFill(fill_type="solid", fgColor=color)
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(path)
    wb.close()


_DIFF_LOWER_IS_BETTER = {"annualized_volatility_diff"}


def _save_value_added_xlsx(df: pd.DataFrame, path: Path) -> None:
    """Save value_added table with conditional formatting on diff columns.

    For return/sharpe/sortino/calmar: positive diff = green (original better).
    For volatility/max_drawdown: positive diff = red (original worse, higher vol/dd).
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, sheet_name="value_added")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        return

    wb  = load_workbook(path)
    ws  = wb.active

    header = [cell.value for cell in ws[1]]

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(header, start=1):
            if not isinstance(col_name, str) or not col_name.endswith("_diff"):
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            val  = cell.value
            if val is None or not isinstance(val, (int, float)):
                continue
            try:
                fval = float(val)
            except (TypeError, ValueError):
                continue
            if np.isnan(fval):
                continue
            lower_is_better = col_name in _DIFF_LOWER_IS_BETTER
            positive_is_good = not lower_is_better
            good = (fval > 0) == positive_is_good
            cell.fill = PatternFill(fill_type="solid", fgColor="E2EFDA" if good else "FCE4D6")

    wb.save(path)
    wb.close()


# ── Scan permutation_test directory ──────────────────────────────────────────

_PERM_PATTERN = re.compile(
    r"^(kmeans|kmedoids_dtw)_ew_inter_markowitz_intra_k(\d+)_seed(\d+)$"
)


def _scan_permutation_dirs() -> dict[tuple[str, int], dict[int, Path]]:
    """
    Returns: { (method_short, k): { seed: Path, ... }, ... }
    method_short is "kmeans" or "kmedoids".
    """
    result: dict[tuple[str, int], dict[int, Path]] = defaultdict(dict)

    if not PERM_TEST_ROOT.exists():
        return result

    for d in sorted(PERM_TEST_ROOT.iterdir()):
        if not d.is_dir():
            continue
        m = _PERM_PATTERN.match(d.name)
        if not m:
            continue
        method_raw, k_str, seed_str = m.group(1), m.group(2), m.group(3)
        method = "kmeans" if method_raw == "kmeans" else "kmedoids"
        result[(method, int(k_str))][int(seed_str)] = d

    return result


# ── Load original strategy ────────────────────────────────────────────────────

def _original_dir(method: str, k: int) -> Path:
    """Return the output directory of the original non-permuted strategy.

    Parameters
    ----------
    method : str
        Clustering method, either ``"kmeans"`` or ``"kmedoids"``.
    k : int
        Number of clusters.

    Returns
    -------
    Path
        Expected output directory for the non-permuted strategy,
        e.g. ``<BACKTESTS_ROOT>/kmeans_ew_inter_markowitz_intra/
        kmeans_ew_inter_markowitz_intra_k5/``.
    """
    prefix = _METHOD_PREFIX[method]
    return BACKTESTS_ROOT / prefix / f"{prefix}_k{k}"


# ── Per-k comparison ──────────────────────────────────────────────────────────

def _build_k_comparison(
    method: str,
    k: int,
    seed_dirs: dict[int, Path],
    out_dir: Path,
) -> dict | None:
    """Build comparison for one (method, k) pair. Returns aggregate row or None.

    Parameters
    ----------
    method : str
        Clustering method, either ``"kmeans"`` or ``"kmedoids"``.
    k : int
        Number of clusters.
    seed_dirs : dict[int, Path]
        Mapping from seed integer to the corresponding permutation
        strategy output directory.
    out_dir : Path
        Directory where comparison CSVs and the colored xlsx are saved.

    Returns
    -------
    dict or None
        Aggregate info dict with keys ``k``, ``original``, and
        ``random_mean`` for use in the all-k tables, or ``None`` if no
        seed data could be loaded.
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    prefix   = _METHOD_PREFIX[method]
    orig_dir = _original_dir(method, k)

    series_dict: dict[str, pd.Series] = {}  # for monthly comparison
    rows: list[dict] = []
    row_colors: dict[str, str] = {}

    # Original strategy
    orig_label = f"{prefix}_k{k}_original"
    if orig_dir.exists():
        daily_orig, monthly_orig = _load_series_from_dir(orig_dir)
        if monthly_orig is not None:
            try:
                row = _metrics_row(orig_label, daily_orig, monthly_orig)
                rows.append(row)
                series_dict[orig_label] = monthly_orig.rename(orig_label)
                row_colors[orig_label]  = _COLOR_ORIGINAL
            except Exception as exc:
                print(f"  [WARN] original k={k} metrics failed: {exc}")
    else:
        print(f"  [WARN] Original strategy not found for {method} k={k}: {orig_dir}")

    # Seed strategies
    seed_monthly: dict[int, pd.Series] = {}
    for seed, seed_dir in sorted(seed_dirs.items()):
        seed_label = f"{prefix}_k{k}_seed{seed:04d}"
        daily_s, monthly_s = _load_series_from_dir(seed_dir)
        if monthly_s is None:
            print(f"  [WARN] No data for {seed_label}, skipping")
            continue
        try:
            row = _metrics_row(seed_label, daily_s, monthly_s)
            rows.append(row)
            series_dict[seed_label] = monthly_s.rename(seed_label)
            row_colors[seed_label]  = _COLOR_SEED
            seed_monthly[seed]       = monthly_s
        except Exception as exc:
            print(f"  [WARN] {seed_label} metrics failed: {exc}")

    if not seed_monthly:
        print(f"  [SKIP] No seed data for {method} k={k}")
        return None

    # random_mean: average of per-seed *metrics* (not metrics of the averaged series).
    # Averaging return series first deflates volatility by 1/sqrt(N), giving a
    # misleadingly low vol for the random baseline.
    random_mean_label = f"{prefix}_k{k}_random_mean"
    seed_rows = [r for r in rows if f"_k{k}_seed" in r["strategy"]]

    mean_row_dict: dict = {"strategy": random_mean_label}
    for col in _METRIC_COLS:
        vals = [r[col] for r in seed_rows if col in r and not np.isnan(float(r[col]))]
        mean_row_dict[col] = float(np.mean(vals)) if vals else np.nan
    mean_row_dict["start_date"] = seed_rows[0].get("start_date", "") if seed_rows else ""
    mean_row_dict["end_date"]   = seed_rows[0].get("end_date",   "") if seed_rows else ""

    rows.append(mean_row_dict)
    row_colors[random_mean_label] = _COLOR_RANDOM_MEAN

    # Averaged series kept only for monthly_returns_comparison / rolling_sharpe charts.
    seed_aligned = pd.DataFrame(seed_monthly).dropna()
    series_dict[random_mean_label] = seed_aligned.mean(axis=1).rename(random_mean_label)

    # Build DataFrames
    summary_df  = pd.DataFrame(rows)
    monthly_df  = pd.concat(series_dict.values(), axis=1).sort_index()
    monthly_df.index.name = "date"
    rolling_df  = _rolling_sharpe(monthly_df, ROLLING_WINDOW)

    # Save
    summary_df.to_csv(out_dir / "summary_metrics.csv", index=False)
    monthly_df.to_csv(out_dir / "monthly_returns_comparison.csv")
    rolling_df.to_csv(out_dir / "rolling_sharpe_comparison.csv")
    _save_xlsx_colored(summary_df, out_dir / "summary_metrics.xlsx", row_colors)

    print(f"  [{method} k={k}] → {out_dir}")

    # Return aggregate info for all_k tables
    orig_row   = next((r for r in rows if r["strategy"] == orig_label), None)
    mean_row   = next((r for r in rows if r["strategy"] == random_mean_label), None)
    return {"k": k, "original": orig_row, "random_mean": mean_row}


# ── all_k tables ──────────────────────────────────────────────────────────────

_METRIC_COLS = [
    "annualized_return",
    "annualized_volatility",
    "sharpe_ratio",
    "sortino_ratio",
    "max_drawdown",
    "calmar_ratio",
]


def _build_value_added(agg_rows: list[dict]) -> pd.DataFrame:
    """Build value_added table: one row per k, original vs random_mean + diff.

    Parameters
    ----------
    agg_rows : list[dict]
        List of aggregate dicts returned by ``_build_k_comparison``,
        each containing ``k``, ``original``, and ``random_mean`` keys.

    Returns
    -------
    pd.DataFrame
        One row per k with columns for original metrics, random_mean
        metrics, and the difference (original minus random_mean).
    """
    records = []
    for agg in sorted(agg_rows, key=lambda x: x["k"]):
        k    = agg["k"]
        orig = agg.get("original") or {}
        mean = agg.get("random_mean") or {}
        row  = {"k": k}
        for col in _METRIC_COLS:
            orig_val = orig.get(col, np.nan)
            mean_val = mean.get(col, np.nan)
            row[f"original_{col}"]    = orig_val
            row[f"random_mean_{col}"] = mean_val
            try:
                diff = float(orig_val) - float(mean_val)
            except (TypeError, ValueError):
                diff = np.nan
            row[f"{col}_diff"] = diff
        records.append(row)
    return pd.DataFrame(records)


def _build_k_comparison_table(agg_rows: list[dict], method: str) -> pd.DataFrame:
    """Build k_comparison: alternating original / random_mean rows sorted by k.

    Parameters
    ----------
    agg_rows : list[dict]
        List of aggregate dicts from ``_build_k_comparison``.
    method : str
        Clustering method used to construct variant label strings.

    Returns
    -------
    pd.DataFrame
        DataFrame with alternating original / random_mean rows for each
        k, columns: ``k``, ``variant``, and all metric columns.
    """
    prefix  = _METHOD_PREFIX[method]
    records = []
    for agg in sorted(agg_rows, key=lambda x: x["k"]):
        k = agg["k"]
        for variant_key, variant_label in [
            ("original",    f"{prefix}_k{k}_original"),
            ("random_mean", f"{prefix}_k{k}_random_mean"),
        ]:
            src = agg.get(variant_key) or {}
            row = {"k": k, "variant": variant_label}
            for col in _METRIC_COLS:
                row[col] = src.get(col, np.nan)
            records.append(row)
    return pd.DataFrame(records)


def _build_all_k(method: str, agg_rows: list[dict], out_dir: Path) -> None:
    """Build and save all-k permutation comparison tables for one method.

    Parameters
    ----------
    method : str
        Clustering method, either ``"kmeans"`` or ``"kmedoids"``.
    agg_rows : list[dict]
        List of aggregate dicts from ``_build_k_comparison``, one per k.
    out_dir : Path
        Output directory where ``value_added`` and ``k_comparison``
        CSV and xlsx files are saved.
    """
    if not agg_rows:
        return
    out_dir.mkdir(parents=True, exist_ok=True)

    va_df   = _build_value_added(agg_rows)
    kc_df   = _build_k_comparison_table(agg_rows, method)

    va_df.to_csv(out_dir / "value_added.csv", index=False)
    kc_df.to_csv(out_dir / "k_comparison.csv", index=False)

    # value_added xlsx (diff cells colored green/red)
    _save_value_added_xlsx(va_df, out_dir / "value_added.xlsx")

    # k_comparison xlsx (original=green, random_mean=orange)
    prefix = _METHOD_PREFIX[method]
    kc_row_colors = {}
    for _, row in kc_df.iterrows():
        v = str(row.get("variant", ""))
        if v.endswith("_original"):
            kc_row_colors[v] = _COLOR_ORIGINAL
        elif v.endswith("_random_mean"):
            kc_row_colors[v] = _COLOR_RANDOM_MEAN

    kc_for_xlsx = kc_df.rename(columns={"variant": "strategy"})
    _save_xlsx_colored(kc_for_xlsx, out_dir / "k_comparison.xlsx", kc_row_colors)

    print(f"  [all_k/{method}] value_added + k_comparison → {out_dir}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    """Parse CLI arguments and build permutation summary outputs."""
    parser = argparse.ArgumentParser(
        description="Build permutation ablation summaries."
    )
    parser.add_argument(
        "--method",
        choices=["kmeans", "kmedoids", "both"],
        default="both",
    )
    parser.add_argument(
        "--k_values",
        nargs="+",
        type=int,
        default=None,
        help="Only process these k values (default: all discovered).",
    )
    parser.add_argument("--universe", default="sp500", choices=UNIVERSE_CHOICES)
    args = parser.parse_args()

    global PERM_TEST_ROOT, BACKTESTS_ROOT, SUMMARY_ROOT
    backtests_dir = get_backtests_dir(args.universe)
    PERM_TEST_ROOT = backtests_dir / "permutation_test"
    BACKTESTS_ROOT = backtests_dir
    SUMMARY_ROOT   = backtests_dir / "summary" / "permutation_summary" / "ew_inter_markowitz_intra"

    methods = ["kmeans", "kmedoids"] if args.method == "both" else [args.method]

    perm_data = _scan_permutation_dirs()
    if not perm_data:
        print(f"[ERROR] No permutation backtest results found under: {PERM_TEST_ROOT}")
        print("Run scripts/run_permutation_backtest.py first.")
        sys.exit(1)

    # Determine which (method, k) pairs to process
    available = {
        (m, k): seed_dirs
        for (m, k), seed_dirs in perm_data.items()
        if m in methods
    }
    if args.k_values:
        available = {(m, k): sd for (m, k), sd in available.items() if k in args.k_values}

    if not available:
        print("[ERROR] No matching (method, k) pairs found.")
        sys.exit(1)

    print(f"Processing {len(available)} (method, k) pairs …")

    # Group by method for the all_k tables
    agg_by_method: dict[str, list[dict]] = defaultdict(list)

    for (method, k), seed_dirs in sorted(available.items()):
        comparison_name = "kmeans_comparison" if method == "kmeans" else "kmedoids_comparison"
        out_dir = SUMMARY_ROOT / f"k{k}" / comparison_name
        agg = _build_k_comparison(method, k, seed_dirs, out_dir)
        if agg is not None:
            agg_by_method[method].append(agg)

    # Build all_k tables per method
    for method, agg_rows in agg_by_method.items():
        all_k_dir = SUMMARY_ROOT / "all_k" / method
        _build_all_k(method, agg_rows, all_k_dir)

    print("\nDone.")
    print(f"Results → {SUMMARY_ROOT}")


if __name__ == "__main__":
    main()
