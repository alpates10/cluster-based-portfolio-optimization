"""
Performance report generator.

--mode classical  (default)
    Reads equal_weight / mean_variance / gmv / cvar strategy folders.
    Writes outputs to data/processed/backtests/summary/classical/

--mode clustering
    Auto-discovers kmeans_equal_weight/kmeans_k*/ and
    kmedoids_dtw_equal_weight/kmedoids_dtw_k*/ sub-folders.
    Writes outputs to data/processed/backtests/summary/clustering/

--mode rolling_validation
    Auto-discovers all sub-folders under
    data/processed/backtests/rolling_validation/.
    Writes outputs to data/processed/backtests/summary/rolling_validation/

--mode all
    Combines classical + clustering + rolling_validation into one table.
    Aligns all series to the start date of the 12m rolling_validation
    strategies (i.e. the latest common start across those series).
    Summary metrics are computed from the aligned monthly data.
    Writes outputs to data/processed/backtests/summary/all/

--mode all_no_rolling
    Combines classical + clustering only (no rolling_validation).
    Uses each strategy's natural start date (no alignment).
    Writes outputs to data/processed/backtests/summary/all_no_rolling/
"""
from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.backtest_summary import (
    build_rolling_sharpe_comparison_table,
    build_monthly_returns_comparison_table,
    build_summary_metrics_table,
    load_strategy_daily_returns,
    load_strategy_monthly_returns_direct,
    save_rolling_sharpe_comparison_table,
    save_monthly_returns_comparison_table,
    save_summary_metrics_excel,
    save_summary_metrics_table,
)
from src.metrics.performance import compute_performance_metrics
from src.paths import get_backtests_dir, UNIVERSE_CHOICES

# ── Constants ─────────────────────────────────────────────────────────────────
CLASSICAL_NAMES = {"equal_weight", "mean_variance", "gmv", "cvar"}
CLUSTERING_PARENTS = ["kmeans_equal_weight", "kmedoids_dtw_equal_weight"]
ADAPTIVE_STRATEGY_NAMES = {"kmeans_adaptive_ew", "kmedoids_dtw_adaptive_ew"}
ROLLING_WINDOW_MONTHS = 12
RISK_FREE_RATE = 0.0
SUMMARY_ROOT = PROJECT_ROOT / "data" / "processed" / "backtests" / "summary"  # overridden in main()
# ─────────────────────────────────────────────────────────────────────────────


def _classical_dirs(backtests_dir: Path) -> list[Path]:
    return [
        p for p in sorted(backtests_dir.iterdir())
        if p.is_dir() and p.name in CLASSICAL_NAMES
    ]


def _clustering_dirs(backtests_dir: Path) -> list[Path]:
    """Auto-discover all clustering strategy directories.

    Scans every directory under *backtests_dir* whose name starts with
    ``kmeans_`` or ``kmedoids_``.

    * If the directory contains sub-directories (e.g. per-k strategy dirs),
      those sub-directories are added individually.
    * If the directory contains no sub-directories it is treated as a flat
      (e.g. adaptive) strategy directory and added directly.

    This covers both the legacy ``kmeans_equal_weight/kmeans_k*/`` layout and
    the newer ``kmeans_markowitz_inter_ew_intra/kmeans_markowitz_*_k*/``
    layout, as well as all flat adaptive strategy directories.
    """
    dirs: list[Path] = []
    for candidate in sorted(backtests_dir.iterdir()):
        if not candidate.is_dir():
            continue
        name = candidate.name
        if not (name.startswith("kmeans_") or name.startswith("kmedoids_")):
            continue

        subdirs = sorted(p for p in candidate.iterdir() if p.is_dir())
        if subdirs:
            dirs.extend(subdirs)
        else:
            # Flat strategy directory (adaptive, or single-strategy folder)
            dirs.append(candidate)

    return dirs


def _rolling_validation_dirs(backtests_dir: Path) -> list[Path]:
    """Auto-discover all strategy directories under rolling_validation/."""
    rv_root = backtests_dir / "rolling_validation"
    if not rv_root.exists():
        return []
    return sorted(
        p for p in rv_root.iterdir()
        if p.is_dir() and not p.name.startswith(".")
    )


# ── Summary metrics from a monthly comparison DataFrame ───────────────────────

def _summary_from_monthly_df(
    monthly_df: pd.DataFrame,
    risk_free_rate: float = 0.0,
    strategy_dirs: list[Path] | None = None,
) -> pd.DataFrame:
    """Compute per-strategy summary metrics from a monthly returns DataFrame.

    Uses periods_per_year=12.  Columns with fewer than 2 valid observations
    are silently skipped.

    n_days and n_months are counted from the actual CSV files within the
    aligned window [monthly_df.index[0], monthly_df.index[-1]].
    """
    dir_by_name: dict[str, Path] = {d.name: d for d in strategy_dirs} if strategy_dirs else {}
    date_start = monthly_df.index[0] if not monthly_df.empty else None
    date_end = monthly_df.index[-1] if not monthly_df.empty else None

    rows: list[dict] = []
    for col in monthly_df.columns:
        series = monthly_df[col].dropna()
        if len(series) < 2:
            print(f"  [SKIP] {col}: fewer than 2 observations after alignment")
            continue
        try:
            metrics = compute_performance_metrics(
                series,
                periods_per_year=12,
                risk_free_rate=risk_free_rate,
            )

            n_days: int | float = float("nan")
            n_months: int | float = float("nan")
            if col in dir_by_name:
                strat_dir = dir_by_name[col]
                try:
                    daily, _ = load_strategy_daily_returns(strat_dir)
                    n_days = int(len(daily.loc[date_start:date_end].dropna()))
                except Exception:
                    pass
                monthly_direct = load_strategy_monthly_returns_direct(strat_dir)
                if monthly_direct is not None:
                    n_months = int(len(monthly_direct.loc[date_start:date_end].dropna()))

            rows.append(
                {
                    "strategy": col,
                    "annualized_return": metrics["annualized_return"],
                    "annualized_volatility": metrics["annualized_volatility"],
                    "sharpe_ratio": metrics["sharpe_ratio"],
                    "max_drawdown": metrics["max_drawdown"],
                    "calmar_ratio": metrics["calmar_ratio"],
                    "n_days": n_days,
                    "n_months": n_months,
                    "start_date": metrics["start_date"],
                    "end_date": metrics["end_date"],
                }
            )
        except Exception as exc:
            print(f"  [SKIP] {col}: metrics computation failed ({exc})")

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows).sort_values("strategy").reset_index(drop=True)
    return df[
        [
            "strategy",
            "annualized_return",
            "annualized_volatility",
            "sharpe_ratio",
            "max_drawdown",
            "calmar_ratio",
            "n_days",
            "n_months",
            "start_date",
            "end_date",
        ]
    ]


# ── Alignment helper ───────────────────────────────────────────────────────────

def _get_alignment_start(monthly_df: pd.DataFrame) -> pd.Timestamp | None:
    """Return the latest first-valid date among 12m rolling validation columns.

    Looks for columns whose name ends with ``_12m``.  Returns None if none
    are present or all are empty.
    """
    rv_12m_cols = [c for c in monthly_df.columns if c.endswith("_12m")]
    if not rv_12m_cols:
        return None
    starts = [monthly_df[c].first_valid_index() for c in rv_12m_cols]
    starts = [s for s in starts if s is not None]
    if not starts:
        return None
    return max(starts)


# ── Generic build-and-save ─────────────────────────────────────────────────────

def _build_and_save(
    backtests_dir: Path,
    strategy_dirs: list[Path],
    out_dir: Path,
    label: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build monthly, summary, and rolling-sharpe tables and save to out_dir.

    Summary metrics are computed from daily returns (preferred) or monthly
    returns as loaded from disk.  Returns (monthly_df, summary_df).
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    monthly_df, monthly_logs = build_monthly_returns_comparison_table(
        backtests_root=backtests_dir,
        strategy_dirs=strategy_dirs,
    )
    if monthly_df.empty:
        print(f"[WARN] {label}: No usable monthly return series found.")
        return pd.DataFrame(), pd.DataFrame()

    summary_df, summary_logs = build_summary_metrics_table(
        backtests_root=backtests_dir,
        risk_free_rate=RISK_FREE_RATE,
        strategy_dirs=strategy_dirs,
    )
    rolling_sharpe_df = build_rolling_sharpe_comparison_table(
        monthly_returns_comparison=monthly_df,
        window_months=ROLLING_WINDOW_MONTHS,
    )

    save_monthly_returns_comparison_table(monthly_df, out_dir / "monthly_returns_comparison.csv")
    save_rolling_sharpe_comparison_table(rolling_sharpe_df, out_dir / "rolling_sharpe_comparison.csv")
    if not summary_df.empty:
        save_summary_metrics_table(summary_df, out_dir / "summary_metrics.csv")
        save_summary_metrics_excel(summary_df, out_dir / "summary_metrics.xlsx")

    print(f"  Monthly comparison  : {out_dir / 'monthly_returns_comparison.csv'}  shape={monthly_df.shape}")
    print(f"  Summary metrics     : {out_dir / 'summary_metrics.csv'}  rows={len(summary_df)}")
    print(f"  Rolling Sharpe      : {out_dir / 'rolling_sharpe_comparison.csv'}  shape={rolling_sharpe_df.shape}")
    print(f"  Strategies: {', '.join(monthly_df.columns)}")

    for line in monthly_logs + summary_logs:
        print(f"  {line}")

    return monthly_df, summary_df


def _build_and_save_aligned(
    backtests_dir: Path,
    strategy_dirs: list[Path],
    out_dir: Path,
    label: str,
    align_start: pd.Timestamp,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build and save tables with all series trimmed to align_start.

    Summary metrics are computed from the aligned monthly data
    (periods_per_year=12) so that every strategy is evaluated over the
    same time interval.
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    monthly_df, monthly_logs = build_monthly_returns_comparison_table(
        backtests_root=backtests_dir,
        strategy_dirs=strategy_dirs,
    )
    if monthly_df.empty:
        print(f"[WARN] {label}: No usable monthly return series found.")
        return pd.DataFrame(), pd.DataFrame()

    # Clip to the alignment start date
    monthly_df = monthly_df.loc[align_start:]

    # Summary from aligned monthly data (periods_per_year=12)
    summary_df = _summary_from_monthly_df(
        monthly_df, risk_free_rate=RISK_FREE_RATE, strategy_dirs=strategy_dirs
    )

    rolling_sharpe_df = build_rolling_sharpe_comparison_table(
        monthly_returns_comparison=monthly_df,
        window_months=ROLLING_WINDOW_MONTHS,
    )

    save_monthly_returns_comparison_table(monthly_df, out_dir / "monthly_returns_comparison.csv")
    save_rolling_sharpe_comparison_table(rolling_sharpe_df, out_dir / "rolling_sharpe_comparison.csv")
    if not summary_df.empty:
        save_summary_metrics_table(summary_df, out_dir / "summary_metrics.csv")
        save_summary_metrics_excel(summary_df, out_dir / "summary_metrics.xlsx")

    print(f"  Aligned from        : {align_start.date()}")
    print(f"  Monthly comparison  : {out_dir / 'monthly_returns_comparison.csv'}  shape={monthly_df.shape}")
    print(f"  Summary metrics     : {out_dir / 'summary_metrics.csv'}  rows={len(summary_df)}")
    print(f"  Rolling Sharpe      : {out_dir / 'rolling_sharpe_comparison.csv'}  shape={rolling_sharpe_df.shape}")
    print(f"  Strategies: {', '.join(monthly_df.columns)}")

    for line in monthly_logs:
        print(f"  {line}")

    return monthly_df, summary_df


# ── All-mode Excel formatting (group colors + legend) ─────────────────────────

def _strategy_group_color(name: str) -> str | None:
    """Return the pastel hex fill color for a strategy, or None if unmatched."""
    if name in {"equal_weight", "mean_variance", "gmv", "cvar"}:
        return "D9D9D9"
    if name.startswith("kmeans_k"):
        return "C9B8E8"
    if name.startswith("kmedoids_dtw_k"):
        return "B8D4F0"
    if name.startswith("kmeans_") and ("markowitz" in name or "ew_inter" in name) and "adaptive" not in name:
        return "B8E8D4"
    if name.startswith("kmedoids_dtw_") and ("markowitz" in name or "ew_inter" in name) and "adaptive" not in name:
        return "F0D4B8"
    if "adaptive" in name:
        return "F0B8C8"
    if "rolling_val" in name:
        return "E8E8B8"
    return None


_LEGEND_ROWS = [
    ("D9D9D9", "Classical baselines",      "EW, Mean-Variance, GMV"),
    ("C9B8E8", "KMeans EW",               "Sabit k, Euclidean clustering"),
    ("B8D4F0", "KMedoids DTW EW",         "Sabit k, DTW clustering"),
    ("B8E8D4", "KMeans Markowitz",        "Sabit k, Markowitz optimizasyonu"),
    ("F0D4B8", "KMedoids DTW Markowitz",  "Sabit k, DTW + Markowitz"),
    ("F0B8C8", "Adaptive silhouette",     "Dinamik k, silhouette bazlı"),
    ("E8E8B8", "Rolling validation",      "Dinamik k, performans bazlı"),
]


def _apply_all_mode_excel_formatting(excel_path: Path, summary_df: pd.DataFrame) -> None:
    """Open the saved xlsx and apply group-based Strategy column colors + legend."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("[WARN] openpyxl not available; skipping all-mode Excel formatting.")
        return

    wb = load_workbook(excel_path)
    ws = wb.active

    # Locate "strategy" column from header row (row 1)
    header_vals = [cell.value for cell in ws[1]]
    if "strategy" not in header_vals:
        print("[WARN] 'strategy' column not found in xlsx; skipping group colors.")
        wb.close()
        return
    strat_col = header_vals.index("strategy") + 1  # 1-based

    # Apply pastel fill to each strategy cell in data rows
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=strat_col)
        if not cell.value:
            continue
        color = _strategy_group_color(str(cell.value))
        if color:
            cell.fill = PatternFill(fill_type="solid", fgColor=color)

    # Legend placement: 2 blank columns after the last data column
    legend_col = len(summary_df.columns) + 3  # +1 for 1-based, +2 for gap

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for offset, label in enumerate(["Renk", "Grup", "Açıklama"]):
        c = ws.cell(row=1, column=legend_col + offset)
        c.value = label
        c.fill = header_fill
        c.font = bold
        c.border = border
        c.alignment = center

    for row_offset, (hex_color, group_name, description) in enumerate(_LEGEND_ROWS, start=2):
        color_cell = ws.cell(row=row_offset, column=legend_col)
        color_cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)
        color_cell.border = border

        name_cell = ws.cell(row=row_offset, column=legend_col + 1)
        name_cell.value = group_name
        name_cell.border = border
        name_cell.alignment = Alignment(vertical="center")

        desc_cell = ws.cell(row=row_offset, column=legend_col + 2)
        desc_cell.value = description
        desc_cell.border = border
        desc_cell.alignment = Alignment(vertical="center")

    wb.save(excel_path)
    wb.close()
    print(f"  All-mode Excel formatting applied: {excel_path}")


# ── Per-mode runners ───────────────────────────────────────────────────────────

def _run_classical(backtests_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    print("\n=== MODE: classical ===")
    strategy_dirs = _classical_dirs(backtests_dir)
    if not strategy_dirs:
        print("[WARN] No classical strategy folders found.")
        return pd.DataFrame(), pd.DataFrame()

    out_dir = SUMMARY_ROOT / "classical"
    monthly_df, summary_df = _build_and_save(backtests_dir, strategy_dirs, out_dir, "classical")

    if not summary_df.empty:
        print("\nSummary metrics:")
        print(summary_df.to_string(index=False))

    return monthly_df, summary_df


def _run_clustering(backtests_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    print("\n=== MODE: clustering ===")
    strategy_dirs = _clustering_dirs(backtests_dir)
    if not strategy_dirs:
        print("[WARN] No clustering sub-folders found — run scripts/run_backtest.py --mode kmeans/kmedoids first.")
        return pd.DataFrame(), pd.DataFrame()

    out_dir = SUMMARY_ROOT / "clustering"
    monthly_df, summary_df = _build_and_save(backtests_dir, strategy_dirs, out_dir, "clustering")

    if not summary_df.empty:
        print("\nCluster comparison table:")
        print(summary_df.to_string(index=False))

    return monthly_df, summary_df


def _run_rolling_validation(backtests_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    print("\n=== MODE: rolling_validation ===")
    strategy_dirs = _rolling_validation_dirs(backtests_dir)
    if not strategy_dirs:
        print(
            "[WARN] No rolling_validation sub-folders found — "
            "run scripts/run_rolling_validation_backtest.py first."
        )
        return pd.DataFrame(), pd.DataFrame()

    out_dir = SUMMARY_ROOT / "rolling_validation"
    monthly_df, summary_df = _build_and_save(
        backtests_dir, strategy_dirs, out_dir, "rolling_validation"
    )

    if not summary_df.empty:
        print("\nRolling validation summary metrics:")
        print(summary_df.to_string(index=False))

    return monthly_df, summary_df


def _run_all(backtests_dir: Path) -> None:
    """classical + clustering + rolling_validation, aligned to 12m start date."""
    print("\n=== MODE: all ===")

    all_dirs = (
        _classical_dirs(backtests_dir)
        + _clustering_dirs(backtests_dir)
        + _rolling_validation_dirs(backtests_dir)
    )
    if not all_dirs:
        print("[WARN] No strategy folders found.")
        return

    out_dir = SUMMARY_ROOT / "all"

    # Peek at monthly data only to determine the alignment date
    monthly_df_full, _ = build_monthly_returns_comparison_table(
        backtests_root=backtests_dir,
        strategy_dirs=all_dirs,
    )
    align_start = _get_alignment_start(monthly_df_full)

    if align_start is None:
        print(
            "[WARN] No 12m rolling_validation series found; "
            "falling back to unaligned build."
        )
        monthly_df, summary_df = _build_and_save(
            backtests_dir, all_dirs, out_dir, "all"
        )
    else:
        monthly_df, summary_df = _build_and_save_aligned(
            backtests_dir, all_dirs, out_dir, "all", align_start
        )

    if not summary_df.empty:
        print(
            "\nCombined summary metrics "
            "(classical + clustering + rolling_validation, aligned):"
        )
        print(summary_df.to_string(index=False))
        _apply_all_mode_excel_formatting(out_dir / "summary_metrics.xlsx", summary_df)


def _run_all_no_rolling(backtests_dir: Path) -> None:
    """classical + clustering only, no alignment (each strategy's natural start)."""
    print("\n=== MODE: all_no_rolling ===")

    all_dirs = _classical_dirs(backtests_dir) + _clustering_dirs(backtests_dir)
    if not all_dirs:
        print("[WARN] No strategy folders found.")
        return

    out_dir = SUMMARY_ROOT / "all_no_rolling"
    monthly_df, summary_df = _build_and_save(
        backtests_dir, all_dirs, out_dir, "all_no_rolling"
    )

    if not summary_df.empty:
        print(
            "\nCombined summary metrics "
            "(classical + clustering, no rolling_validation):"
        )
        print(summary_df.to_string(index=False))


# ── Legacy cleanup ─────────────────────────────────────────────────────────────

def _delete_legacy_files(backtests_dir: Path) -> None:
    """Remove old flat summary files from backtests/ root (now kept in summary/)."""
    legacy = [
        "summary_metrics.csv",
        "summary_metrics.xlsx",
        "monthly_returns_comparison.csv",
        "rolling_sharpe_comparison.csv",
    ]
    for name in legacy:
        p = backtests_dir / name
        if p.exists():
            p.unlink()
            print(f"Removed legacy file: {p}")


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate performance reports.")
    parser.add_argument(
        "--mode",
        choices=[
            "classical",
            "clustering",
            "rolling_validation",
            "all",
            "all_no_rolling",
        ],
        default="classical",
        help=(
            "classical (default) → equal_weight/mean_variance/gmv/cvar → summary/classical/; "
            "clustering → kmeans_k*/kmedoids_dtw_k* → summary/clustering/; "
            "rolling_validation → rolling_validation/* → summary/rolling_validation/; "
            "all → classical+clustering+rolling_validation aligned to 12m start → summary/all/; "
            "all_no_rolling → classical+clustering, natural start dates → summary/all_no_rolling/"
        ),
    )
    parser.add_argument("--universe", default="sp500", choices=UNIVERSE_CHOICES)
    args = parser.parse_args()

    global SUMMARY_ROOT
    backtests_dir = get_backtests_dir(args.universe)
    SUMMARY_ROOT = backtests_dir / "summary"
    _delete_legacy_files(backtests_dir)

    if args.mode == "classical":
        _run_classical(backtests_dir)
    elif args.mode == "clustering":
        _run_clustering(backtests_dir)
    elif args.mode == "rolling_validation":
        _run_rolling_validation(backtests_dir)
    elif args.mode == "all":
        _run_all(backtests_dir)
    elif args.mode == "all_no_rolling":
        _run_all_no_rolling(backtests_dir)


if __name__ == "__main__":
    main()
